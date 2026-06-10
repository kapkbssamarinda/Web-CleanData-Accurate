import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GL Cleaner – Accurate 5",
    page_icon="📒",
    layout="wide",
)

st.markdown("""
<style>
    .stApp { background-color: #0f1b2d; color: #e8dcc8; }
    .main-title { color: #c9a84c; font-size: 1.8rem; font-weight: 700; margin-bottom: 0; }
    .sub-title  { color: #8a9bb5; font-size: 0.95rem; margin-top: 4px; margin-bottom: 1.5rem; }
    .stat-card  {
        background: #1a2d4a; border: 1px solid #2d4a6e;
        border-radius: 8px; padding: 16px 20px; text-align: center;
    }
    .stat-num { color: #c9a84c; font-size: 1.7rem; font-weight: 700; line-height: 1; }
    .stat-lbl { color: #8a9bb5; font-size: 0.8rem; margin-top: 4px; }
    .info-box {
        background: #132238; border-left: 3px solid #c9a84c;
        border-radius: 4px; padding: 10px 14px;
        color: #c8d6e8; font-size: 0.88rem; margin: 8px 0;
    }
    .warn-box {
        background: #2a1a0a; border-left: 3px solid #e07b2a;
        border-radius: 4px; padding: 10px 14px;
        color: #e8c8a0; font-size: 0.88rem; margin: 8px 0;
    }
    div[data-testid="stDataFrame"] { border: 1px solid #2d4a6e; border-radius: 6px; }
    .stButton > button {
        background: #c9a84c; color: #0f1b2d;
        border: none; font-weight: 600; border-radius: 6px;
        padding: 0.5rem 2rem; font-size: 1rem;
    }
    .stButton > button:hover { background: #e0bf6a; color: #0f1b2d; }
    .stDownloadButton > button {
        background: #1a6e3a; color: #d0f0e0;
        border: none; font-weight: 600; border-radius: 6px;
        padding: 0.5rem 2rem; font-size: 1rem; width: 100%;
    }
    .stDownloadButton > button:hover { background: #228048; }
    [data-testid="stFileUploader"] {
        background: #132238; border: 2px dashed #2d4a6e;
        border-radius: 8px; padding: 10px;
    }
    .section-head {
        color: #c9a84c; font-size: 1rem; font-weight: 600;
        border-bottom: 1px solid #2d4a6e; padding-bottom: 6px;
        margin: 1.2rem 0 0.8rem;
    }
    .badge {
        display: inline-block; padding: 2px 8px; border-radius: 10px;
        font-size: 0.75rem; font-weight: 600;
    }
    .badge-ok  { background: #0e3320; color: #50d89a; }
    .badge-err { background: #3a1010; color: #f07070; }
</style>
""", unsafe_allow_html=True)

# ─── Constants ─────────────────────────────────────────────────────────────────
OUTPUT_COLS = ["Tanggal", "COA", "Nama Akun", "Sumber", "No. Sumber", "Keterangan", "Debit", "Kredit", "Balance"]
NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def get_cells(row):
    """Return {col_index: value} for a SpreadsheetML row, handling ss:Index."""
    result = {}
    col = 0
    for cell in row.findall("ss:Cell", NS):
        idx = cell.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
        if idx:
            col = int(idx) - 1
        data = cell.find("ss:Data", NS)
        val = (data.text or "").strip() if data is not None else ""
        if val:
            result[col] = val
        col += 1
    return result


def is_date(s):
    return bool(re.match(r"^\d{1,2}\s+\w+\s+\d{4}$", s.strip()))


def is_coa(s):
    return bool(re.match(r"^\d{3,}[\.\d\-]*$", s.strip()))


def clean_num(s):
    """Strip thousand separators (dots in ID format), normalise to plain string."""
    if not s or s == "0,00":
        return "0"
    # Remove thousand separators (dots), keep comma as decimal
    s = re.sub(r"\.", "", s)
    s = s.replace(",", ".")
    try:
        return str(float(s))
    except Exception:
        return s


# ─── Auto-detect column layout ────────────────────────────────────────────────

def detect_layout(rows):
    """
    Scan the first ~80 rows to auto-detect column positions.
    Returns dict with keys: coa, nama_akun, saldo_val, saldo_dc,
                            tanggal, sumber, no_sumber, keterangan,
                            debit, kredit, balance
    """
    # Gather all non-empty (col, value) from first 80 rows
    sample = []
    for row in rows[:80]:
        sample.append(get_cells(row))

    # --- Detect transaction rows first ---
    # A transaction row has a date at some column, and numeric-looking values later
    date_cols = {}
    for cv in sample:
        for c, v in cv.items():
            if is_date(v):
                date_cols[c] = date_cols.get(c, 0) + 1

    if not date_cols:
        return None

    tanggal_col = max(date_cols, key=date_cols.get)

    # For transaction rows, find columns relative to tanggal
    # Typical pattern: sumber ~+5, no_sumber ~+10, keterangan ~+16, debit/kredit/balance after
    # We'll use heuristics: after tanggal col, find "Bukti Jurnal"/"BJ"/"KK" etc patterns
    tx_rows = [cv for cv in sample if tanggal_col in cv and is_date(cv[tanggal_col])]

    # Sumber: col with "Bukti Jurnal" / "BJ" / journal-type strings
    sumber_candidates = {}
    no_sumber_candidates = {}
    ket_candidates = {}
    num_candidates = {}   # columns with numeric-looking values

    for cv in tx_rows:
        for c, v in cv.items():
            if c == tanggal_col:
                continue
            if re.match(r"^[A-Z]{1,8}[-/]\d+", v):  # No.Sumber: KK-202501001, BP/2025/001
                no_sumber_candidates[c] = no_sumber_candidates.get(c, 0) + 1
            elif re.match(r"^(Bukti Jurnal|Bukti Kas|Bukti Bank|Jurnal Umum|Memo|Penyesuaian)$", v, re.I):
                sumber_candidates[c] = sumber_candidates.get(c, 0) + 1
            elif re.match(r"^(Bukti Jurnal|BJ|Jurnal)", v, re.I) and len(v) <= 20:
                sumber_candidates[c] = sumber_candidates.get(c, 0) + 1
            elif re.match(r"^\d[\d\.,]+$", v):         # numeric string
                num_candidates[c] = num_candidates.get(c, 0) + 1

    # Keterangan: free-text column between no_sumber and debit
    # After we have sumber/no_sumber positions, keterangan is next notable text col
    sumber_col    = max(sumber_candidates,    key=sumber_candidates.get)    if sumber_candidates    else tanggal_col + 5
    no_sumber_col = max(no_sumber_candidates, key=no_sumber_candidates.get) if no_sumber_candidates else sumber_col + 5

    # Keterangan: text columns after no_sumber, before numeric cols
    num_cols_sorted = sorted(num_candidates.keys())
    first_num_col = num_cols_sorted[0] if num_cols_sorted else no_sumber_col + 10
    ket_text_cols = {}
    for cv in tx_rows:
        for c, v in cv.items():
            if c > no_sumber_col and c < first_num_col:
                if not re.match(r"^\d[\d\.,]+$", v):
                    ket_text_cols[c] = ket_text_cols.get(c, 0) + 1
    ket_col = max(ket_text_cols, key=ket_text_cols.get) if ket_text_cols else no_sumber_col + 6

    # Numeric columns: identify debit, kredit, balance by position + frequency
    # Balance usually contains "(Dr)" / "(Cr)" prefix
    balance_candidates = {}
    for cv in tx_rows:
        for c, v in cv.items():
            if re.search(r"\(Dr\)|\(Cr\)", v):
                balance_candidates[c] = balance_candidates.get(c, 0) + 1

    balance_col = max(balance_candidates, key=balance_candidates.get) if balance_candidates else None

    # Among remaining numeric cols, debit & kredit are the two most frequent
    # that are NOT balance_col, sorted by column position (debit before kredit)
    remaining_nums = {c: cnt for c, cnt in num_candidates.items() if c != balance_col}
    sorted_num = sorted(remaining_nums, key=lambda c: (-remaining_nums[c], c))
    top_num = sorted(sorted_num[:4])  # take up to 4 highest-frequency, sort by col position

    # Debit = lower col index, Kredit = higher
    if len(top_num) >= 2:
        debit_col, kredit_col = top_num[0], top_num[1]
    elif len(top_num) == 1:
        debit_col = kredit_col = top_num[0]
    else:
        debit_col = ket_col + 8
        kredit_col = ket_col + 12

    # --- Detect account-header rows ---
    # Account header: one cell with COA pattern, another with name, optionally saldo
    coa_candidates = {}
    nama_candidates = {}
    saldo_val_candidates = {}
    saldo_dc_candidates = {}

    for cv in sample:
        # Skip if this looks like a transaction row
        if tanggal_col in cv and is_date(cv.get(tanggal_col, "")):
            continue
        for c, v in cv.items():
            if is_coa(v):
                coa_candidates[c] = coa_candidates.get(c, 0) + 1
            elif re.match(r"^(Dr|Cr)$", v):
                saldo_dc_candidates[c] = saldo_dc_candidates.get(c, 0) + 1

    coa_col      = max(coa_candidates,      key=coa_candidates.get)      if coa_candidates      else 1
    saldo_dc_col = max(saldo_dc_candidates, key=saldo_dc_candidates.get) if saldo_dc_candidates else None

    # Nama Akun: text col near coa_col in header rows, but not coa itself
    # and not a known numeric or date col
    hdr_rows = [cv for cv in sample if coa_col in cv and is_coa(cv.get(coa_col, ""))]
    nama_text_cols = {}
    for cv in hdr_rows:
        for c, v in cv.items():
            if c != coa_col and not is_coa(v) and not re.match(r"^\d[\d\.,]+$", v) and not re.match(r"^(Dr|Cr)$", v):
                nama_text_cols[c] = nama_text_cols.get(c, 0) + 1
    nama_col = max(nama_text_cols, key=nama_text_cols.get) if nama_text_cols else coa_col + 7

    # Saldo val: numeric col in header rows
    saldo_num_cols = {}
    for cv in hdr_rows:
        for c, v in cv.items():
            # Exclude coa_col itself and any column whose value looks like a COA
            if c == coa_col:
                continue
            if re.match(r"^\d[\d\.,]+$", v) and not is_coa(v):
                saldo_num_cols[c] = saldo_num_cols.get(c, 0) + 1
    saldo_val_col = max(saldo_num_cols, key=saldo_num_cols.get) if saldo_num_cols else None

    return {
        "coa":        coa_col,
        "nama_akun":  nama_col,
        "saldo_val":  saldo_val_col,
        "saldo_dc":   saldo_dc_col,
        "tanggal":    tanggal_col,
        "sumber":     sumber_col,
        "no_sumber":  no_sumber_col,
        "keterangan": ket_col,
        "debit":      debit_col,
        "kredit":     kredit_col,
        "balance":    balance_col,
    }


# ─── Parser ───────────────────────────────────────────────────────────────────

def parse_spreadsheetml(file_bytes):
    """Parse SpreadsheetML XML (Accurate 5 .xls export format)."""
    tree = ET.parse(io.BytesIO(file_bytes))
    root = tree.getroot()
    sheets = root.findall(".//ss:Worksheet", NS)
    if not sheets:
        raise ValueError("Tidak ditemukan worksheet dalam file.")
    return sheets


def extract_records(rows, layout, keep_raw_balance=True):
    """
    Walk all rows, detect account headers and transaction lines,
    return list of record dicts.
    """
    records = []
    current_coa = ""
    current_nama = ""

    L = layout  # shorthand

    for row in rows:
        cv = get_cells(row)
        if not cv:
            continue

        c1 = cv.get(L["coa"], "")

        # ── Account header row ─────────────────────────────────────
        if c1 and is_coa(c1) and cv.get(L["nama_akun"]):
            current_coa  = c1
            current_nama = cv.get(L["nama_akun"], "")

            saldo_raw = cv.get(L["saldo_val"], "0,00") if L["saldo_val"] else "0,00"
            dc = cv.get(L["saldo_dc"], "") if L["saldo_dc"] else ""

            if dc == "Dr":
                debit_sa, kredit_sa = saldo_raw, "0,00"
            elif dc == "Cr":
                debit_sa, kredit_sa = "0,00", saldo_raw
            else:
                debit_sa, kredit_sa = saldo_raw, "0,00"

            balance_sa = saldo_raw  # plain number, no Dr/Cr prefix

            records.append({
                "Tanggal":    "Saldo Awal",
                "COA":        current_coa,
                "Nama Akun":  current_nama,
                "Sumber":     "",
                "No. Sumber": "",
                "Keterangan": "Saldo Awal",
                "Debit":      debit_sa,
                "Kredit":     kredit_sa,
                "Balance":    balance_sa,
                "_is_saldo":  True,
            })
            continue

        # ── Transaction row ────────────────────────────────────────
        tanggal = cv.get(L["tanggal"], "")
        if tanggal and is_date(tanggal):
            balance_raw = cv.get(L["balance"], "") if L["balance"] else ""
            balance_val = re.sub(r"\(Dr\)\s*|\(Cr\)\s*", "", balance_raw).strip()
            records.append({
                "Tanggal":    tanggal,
                "COA":        current_coa,
                "Nama Akun":  current_nama,
                "Sumber":     cv.get(L["sumber"], ""),
                "No. Sumber": cv.get(L["no_sumber"], ""),
                "Keterangan": cv.get(L["keterangan"], ""),
                "Debit":      cv.get(L["debit"], "0,00"),
                "Kredit":     cv.get(L["kredit"], "0,00"),
                "Balance":    balance_val,
                "_is_saldo":  False,
            })

    return records


# ─── XLSX export ──────────────────────────────────────────────────────────────

def build_xlsx(records_df):
    """
    Build a clean, filter-friendly XLSX.
    Uses iloc positional access — avoids itertuples mangling column names
    that contain spaces or dots (e.g. "Nama Akun", "No. Sumber").
    """
    import math
    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Besar"

    hdr_font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="0F3460")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell_border = Border(bottom=Side(style="thin", color="EBEBEB"))

    for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    saldo_fill  = PatternFill("solid", fgColor="EBF5FF")
    right_cols  = {"Debit", "Kredit", "Balance"}

    # Build positional index maps — safe against spaces/dots in column names
    all_cols      = list(records_df.columns)
    out_positions = [all_cols.index(c) for c in OUTPUT_COLS]
    saldo_pos     = all_cols.index("_is_saldo")

    def safe_str(v):
        if v is None: return ""
        if isinstance(v, float) and math.isnan(v): return ""
        return str(v)

    for df_idx in range(len(records_df)):
        row_vals = records_df.iloc[df_idx]
        is_saldo = bool(row_vals.iloc[saldo_pos])
        xlsx_row = df_idx + 2
        for col_idx, pos in enumerate(out_positions, 1):
            col_name = OUTPUT_COLS[col_idx - 1]
            val      = safe_str(row_vals.iloc[pos])
            cell     = ws.cell(row=xlsx_row, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(
                horizontal="right" if col_name in right_cols else "left",
                vertical="center",
            )
            if is_saldo:
                cell.fill = saldo_fill
            cell.border = cell_border

    col_max = {col: len(col) for col in OUTPUT_COLS}
    for df_idx in range(len(records_df)):
        row_vals = records_df.iloc[df_idx]
        for col_name, pos in zip(OUTPUT_COLS, out_positions):
            val = safe_str(row_vals.iloc[pos])
            col_max[col_name] = min(max(col_max[col_name], len(val)), 60)

    for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_max[col_name] + 3

    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLS))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── UI ───────────────────────────────────────────────────────────────────────

st.markdown('<p class="main-title">📒 GL Cleaner — Accurate 5</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">Bersihkan & ekspor Buku Besar Rinci dari Accurate 5 Desktop menjadi Excel yang bisa di-filter</p>', unsafe_allow_html=True)

# ── Upload ─────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload file ekspor Buku Besar (.xls / .xlsx dari Accurate 5)",
    type=["xls", "xlsx"],
    help="File SpreadsheetML XML yang dihasilkan Accurate 5 Desktop",
)

if not uploaded:
    st.markdown("""
    <div class="info-box">
    ℹ️ <b>Cara pakai:</b><br>
    1. Di Accurate 5 Desktop, buka <b>Buku Besar Rinci</b> → klik <b>Ekspor ke Excel</b><br>
    2. Upload file .xls hasil ekspor di atas<br>
    3. Preview data, lalu klik <b>Download XLSX</b>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Parse ──────────────────────────────────────────────────────────────────────
with st.spinner("Membaca dan mendeteksi struktur file..."):
    try:
        file_bytes = uploaded.read()

        # Check if SpreadsheetML
        header_bytes = file_bytes[:200]
        is_xml = b"<?xml" in header_bytes or b"<Workbook" in header_bytes

        if not is_xml:
            st.error("Format file tidak dikenali. Pastikan file adalah hasil ekspor Accurate 5 (SpreadsheetML).")
            st.stop()

        sheets = parse_spreadsheetml(file_bytes)

        # Use first sheet
        sheet = sheets[0]
        sheet_name = sheet.get("{urn:schemas-microsoft-com:office:spreadsheet}Name", "Sheet 1")
        table = sheet.find("ss:Table", NS)
        rows  = table.findall("ss:Row", NS)

        # Auto-detect layout
        layout = detect_layout(rows)
        if layout is None:
            st.error("Tidak dapat mendeteksi struktur kolom. Pastikan file berisi data transaksi Buku Besar.")
            st.stop()

        # Extract records
        records = extract_records(rows, layout)

    except ET.ParseError as e:
        st.error(f"File XML tidak valid: {e}")
        st.stop()
    except Exception as e:
        st.error(f"Gagal membaca file: {e}")
        st.stop()

# ── Stats ──────────────────────────────────────────────────────────────────────
if not records:
    st.warning("Tidak ada data transaksi yang berhasil diekstrak.")
    st.stop()

df_all = pd.DataFrame(records)
df_out = df_all[OUTPUT_COLS + ["_is_saldo"]].copy()

n_saldo = df_out["_is_saldo"].sum()
n_tx    = len(df_out) - n_saldo
n_akun  = df_out[df_out["_is_saldo"]]["COA"].nunique()
n_sheet = len(sheets)

c1, c2, c3, c4 = st.columns(4)
for col, num, label in [
    (c1, n_tx,    "Transaksi"),
    (c2, n_akun,  "Akun (COA)"),
    (c3, n_saldo, "Baris Saldo Awal"),
    (c4, len(rows), "Total Baris File"),
]:
    col.markdown(f'<div class="stat-card"><div class="stat-num">{num:,}</div><div class="stat-lbl">{label}</div></div>', unsafe_allow_html=True)

# ── Layout info ────────────────────────────────────────────────────────────────
with st.expander("🔍 Hasil deteksi kolom otomatis", expanded=False):
    labels = {
        "coa": "COA", "nama_akun": "Nama Akun", "saldo_val": "Saldo Awal",
        "saldo_dc": "Dr/Cr Saldo", "tanggal": "Tanggal", "sumber": "Sumber",
        "no_sumber": "No. Sumber", "keterangan": "Keterangan",
        "debit": "Debit", "kredit": "Kredit", "balance": "Balance",
    }
    rows_det = []
    for key, label in labels.items():
        col_idx = layout.get(key)
        col_letter = get_column_letter(col_idx + 1) if col_idx is not None else "—"
        rows_det.append({"Kolom Output": label, "Posisi Kolom (0-based)": col_idx, "Huruf Kolom": col_letter})
    st.dataframe(pd.DataFrame(rows_det), use_container_width=True, hide_index=True)

# ── Filter sidebar ─────────────────────────────────────────────────────────────
st.sidebar.markdown("## Filter Data")

all_coa = sorted(df_out["COA"].dropna().unique().tolist())
sel_coa = st.sidebar.multiselect("COA / Akun", all_coa, default=[])

all_sumber = sorted(df_out["Sumber"].dropna().replace("", pd.NA).dropna().unique().tolist())
sel_sumber = st.sidebar.multiselect("Sumber", all_sumber, default=[])

search_ket = st.sidebar.text_input("Cari Keterangan", "")
show_saldo = st.sidebar.checkbox("Tampilkan baris Saldo Awal", value=True)

# Apply filters
df_view = df_out.copy()
if sel_coa:
    df_view = df_view[df_view["COA"].isin(sel_coa)]
if sel_sumber:
    df_view = df_view[df_view["Sumber"].isin(sel_sumber)]
if search_ket:
    df_view = df_view[df_view["Keterangan"].str.contains(search_ket, case=False, na=False)]
if not show_saldo:
    df_view = df_view[~df_view["_is_saldo"]]

# ── Preview ────────────────────────────────────────────────────────────────────
st.markdown(f'<p class="section-head">Preview — {len(df_view):,} baris</p>', unsafe_allow_html=True)

st.dataframe(
    df_view[OUTPUT_COLS].head(500),
    use_container_width=True,
    hide_index=True,
    height=420,
)

if len(df_view) > 500:
    st.caption(f"⚠️ Preview dibatasi 500 baris. File XLSX akan berisi semua {len(df_view):,} baris.")

# ── Download ───────────────────────────────────────────────────────────────────
st.markdown('<p class="section-head">Export</p>', unsafe_allow_html=True)

export_choice = st.radio(
    "Data yang diekspor:",
    ["Semua data (tanpa filter)", "Data sesuai filter aktif"],
    horizontal=True,
)

df_export = df_out if export_choice.startswith("Semua") else df_view

with st.spinner("Menyiapkan file XLSX..."):
    xlsx_bytes = build_xlsx(df_export)

fname_base = uploaded.name.rsplit(".", 1)[0]
st.download_button(
    label=f"⬇️  Download XLSX  ({len(df_export):,} baris)",
    data=xlsx_bytes,
    file_name=f"{fname_base}_clean.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.markdown("""
<div class="info-box" style="margin-top:1rem;">
ℹ️ Output XLSX: header beku di baris 1 · AutoFilter aktif · lebar kolom otomatis · baris Saldo Awal diberi warna biru muda · tanpa formula · siap di-filter / pivot
</div>
""", unsafe_allow_html=True)